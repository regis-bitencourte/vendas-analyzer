# reports.py
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def generate_excel_report(analysis_data: dict) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # --- ABA 1: RESUMO ---
    ws = wb.active
    ws.title = "Resumo Executivo"
    
    metrics = [
        ("RESUMO FINANCEIRO", ""),
        ("Pedidos Pagos", analysis_data.get('paid_count', 0)),
        ("Total Recebido (R$)", analysis_data.get('total_received', 0)),
        ("Lucro Líquido (R$)", analysis_data.get('net_profit', 0)),
        ("Ticket Médio (R$)", analysis_data.get('avg_ticket', 0)),
        ("CAC - Custo Aquisição (R$)", analysis_data.get('cac', 0)),
        ("LTV - Valor do Cliente (R$)", analysis_data.get('ltv', 0)),
    ]
    
    row = 1
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value if value != "" else ""
        if value == "": ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

    # --- ABA 2: FRETES ---
    ws = wb.create_sheet("Fretes")
    ws.append(["MÉTODO", "QTD PEDIDOS", "VALOR TOTAL COBRADO (R$)"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    
    f_stats = analysis_data.get('shipping_stats', {})
    ws.append(["Frete Grátis", f_stats.get('gratis', {}).get('count', 0), "0.00"])
    ws.append(["PAC", f_stats.get('pac', {}).get('count', 0), f_stats.get('pac', {}).get('value', 0)])
    ws.append(["SEDEX", f_stats.get('sedex', {}).get('count', 0), f_stats.get('sedex', {}).get('value', 0)])
    ws.append(["Transportadora", f_stats.get('transportadora', {}).get('count', 0), f_stats.get('transportadora', {}).get('value', 0)])

    # --- ABA 3: ABC ---
    ws = wb.create_sheet("Curva ABC")
    ws.append(["PRODUTO", "RECEITA GERADA (R$)", "CURVA"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    abc_data = analysis_data.get('abc_curve', {})
    for item in abc_data.get('A', []): ws.append([item['product'], item['revenue'], 'A (80%)'])
    for item in abc_data.get('B', []): ws.append([item['product'], item['revenue'], 'B (15%)'])
    for item in abc_data.get('C', []): ws.append([item['product'], item['revenue'], 'C (5%)'])

    # --- ABA 4: CIDADES ---
    ws = wb.create_sheet("Geográfico")
    ws.append(["ESTADO", "RECEITA (R$)", "PEDIDOS"])
    for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
    geo = analysis_data.get('geographic', {})
    for state in geo.get('top_states', []): ws.append([state[0], state[1], state[2]])
    ws.append(["", "", ""])
    ws.append(["CIDADE", "RECEITA (R$)", "PEDIDOS"])
    for cell in ws[ws.max_row]: cell.fill, cell.font = header_fill, header_font
    for city in geo.get('top_cities', []): ws.append([city[0], city[1], city[2]])

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf_report(analysis_data: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1f77b4'), alignment=1, spaceAfter=20)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1f77b4'), spaceBefore=15, spaceAfter=10)

    elements.append(Paragraph(f"Relatório Executivo de Vendas", title_style))
    
    # 1. RESUMO
    elements.append(Paragraph("MÉTRICAS PRINCIPAIS E MARKETING", heading_style))
    data = [
        ['Pedidos Pagos', str(analysis_data.get('paid_count', 0))],
        ['Receita Total', f"R$ {analysis_data.get('total_received', 0):,.2f}"],
        ['Lucro Líquido', f"R$ {analysis_data.get('net_profit', 0):,.2f}"],
        ['CAC (Custo Aquisição)', f"R$ {analysis_data.get('cac', 0):,.2f}"],
        ['LTV (Valor do Cliente)', f"R$ {analysis_data.get('ltv', 0):,.2f}"]
    ]
    t = Table(data, colWidths=[3*inch, 2*inch])
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('BOTTOMPADDING', (0, 0), (-1, -1), 10)]))
    elements.append(t)

    # 2. FRETES
    elements.append(Paragraph("DETALHAMENTO DE FRETES", heading_style))
    f_stats = analysis_data.get('shipping_stats', {})
    data = [
        ['Método', 'Qtd. Pedidos', 'Valor Arrecadado (R$)'],
        ['Frete Grátis', str(f_stats.get('gratis', {}).get('count', 0)), "R$ 0.00"],
        ['PAC', str(f_stats.get('pac', {}).get('count', 0)), f"R$ {f_stats.get('pac', {}).get('value', 0):,.2f}"],
        ['SEDEX', str(f_stats.get('sedex', {}).get('count', 0)), f"R$ {f_stats.get('sedex', {}).get('value', 0):,.2f}"],
        ['Transportadora', str(f_stats.get('transportadora', {}).get('count', 0)), f"R$ {f_stats.get('transportadora', {}).get('value', 0):,.2f}"]
    ]
    t = Table(data, colWidths=[2*inch, 1.5*inch, 2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(t)

    # 3. PAGAMENTOS
    elements.append(Paragraph("TAXAS POR PAGAMENTO", heading_style))
    data = [['Método', 'Pedidos', 'Taxas (R$)']]
    for method, stats in analysis_data.get('payment_stats', {}).items():
        if stats['count'] > 0:
            data.append([method.upper(), str(stats['count']), f"R$ {stats.get('tax_amount', 0):,.2f}"])
    t = Table(data, colWidths=[2*inch, 1.5*inch, 2*inch])
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(t)

    # 4. TOP CIDADES E ESTADOS
    geo = analysis_data.get('geographic', {})
    if geo.get('top_states'):
        elements.append(Paragraph("TOP ESTADOS", heading_style))
        data = [['Estado', 'Receita', 'Pedidos']]
        for state in geo['top_states'][:5]: data.append([str(state[0]), f"R$ {state[1]:,.2f}", str(state[2])])
        t = Table(data, colWidths=[2*inch, 2*inch, 1*inch])
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
        elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()