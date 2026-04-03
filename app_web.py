# -*- coding: utf-8 -*-
"""
Analisador de Vendas Web
Aplicação web para análise de dados de vendas com geração de relatórios em PDF.
"""

import io
import json
import os
import re
import logging
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple

import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuração da página Streamlit
st.set_page_config(
    page_title="Analisador de Vendas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Tema e estilos
st.markdown("""
<style>
    .main {
        padding: 0rem 1rem;
    }
    .metric-container {
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        margin: 10px 0;
    }
    h1 {
        color: #1f77b4;
        text-align: center;
    }
    .section-title {
        color: #1f77b4;
        font-size: 1.3em;
        font-weight: bold;
        margin-top: 20px;
    }
</style>
""", unsafe_allow_html=True)

# Constantes
DEFAULT_STORE_NAME = "OnFight"
DEFAULT_OVERSIZED_COST = 30.00
DEFAULT_PIX_TAX = 0.00
DEFAULT_BOLETO_TAX_BRL = 1.50
DEFAULT_GATEWAY_TAX_BRL = 1.50

# Arquivo de histórico
HISTORY_FILE = "analise_history.json"

# Categorias padrão (fallback)
DEFAULT_CATEGORIAS_CONFIG = {
    "Oversized": ["oversized"],
    "Kit Dryfit": ["kit camiseta dryfit", "kit dryfit", "camiseta dryfit com short", "short dry fit"],
    "Camiseta Infantil": ["camiseta infantil", "camisetas infantil", "infantil jiu-jitsu"],
    "Bermuda Dry Fit": ["bermuda masculina dry fit", "bermuda dry fit", "conforto e performance"],
    "Short 2 em 1": ["short", "2 em 1", "2em1"],
    "Dryfit": ["dryfit", "dry fit"],
    "Moletom": ["moletom", "hoodie"],
    "Calça": ["calça", "calca", "pants"],
    "Combo": ["combo", "kit"]
}


class VendasAnalyzerWeb:
    """Analisador de vendas para interface web."""

    def __init__(self, categorias_config: Dict[str, list] = None):
        """Inicializa o analisador."""
        self.categorias_config = categorias_config or DEFAULT_CATEGORIAS_CONFIG.copy()

    def _identify_category(self, product_name: str) -> str:
        """
        Identifica a categoria do produto baseado em palavras-chave.
        """
        product_name_lower = str(product_name).lower().strip()
        
        for category, keywords in self.categorias_config.items():
            for keyword in keywords:
                if keyword.lower() in product_name_lower:
                    return category
        
        return "Outros"

    def add_category(self, category_name: str, keywords: list) -> bool:
        """Adiciona uma nova categoria."""
        if not category_name.strip():
            return False
        clean_keywords = [kw.strip().lower() for kw in keywords if kw.strip()]
        if not clean_keywords:
            return False
        self.categorias_config[category_name.strip()] = clean_keywords
        return True

    def remove_category(self, category_name: str) -> bool:
        """Remove uma categoria."""
        if category_name in self.categorias_config:
            del self.categorias_config[category_name]
            return True
        return False

    def update_category(self, old_name: str, new_name: str, keywords: list) -> bool:
        """Atualiza uma categoria existente."""
        if old_name not in self.categorias_config:
            return False
        if old_name != new_name:
            del self.categorias_config[old_name]
        return self.add_category(new_name, keywords)

    def get_categories_list(self) -> list:
        """Retorna lista de categorias disponíveis."""
        return [cat for cat in self.categorias_config.keys() if cat != "Outros"]

    @staticmethod
    def _parse_float(value: str) -> float:
        """Converte string para float."""
        try:
            return float(str(value).strip().replace(',', '.'))
        except ValueError:
            raise ValueError(f"Valor inválido: '{value}'")

    def _identify_payment_method(self, row: pd.Series) -> str:
        """
        Identifica o método de pagamento avaliando várias colunas.
        """
        payment_str = str(row.get('Payment Method', '')) + " " + str(row.get('Tags', '')) + " " + str(row.get('Note Attributes', ''))
        payment_lower = payment_str.lower()
        
        if 'pix' in payment_lower:
            return 'pix'
        elif 'boleto' in payment_lower:
            return 'boleto'
        elif 'cartao' in payment_lower or 'cartão' in payment_lower or 'credit' in payment_lower:
            return 'cartao'
        else:
            return 'outro'

    def _identify_installments(self, row: pd.Series) -> int:
        """
        Extrai o número de parcelas do pedido procurado em várias colunas do CSV da Shopify.
        """
        payment_str = str(row.get('Payment Method', '')) + " " + str(row.get('Tags', '')) + " " + str(row.get('Note Attributes', '')) + " " + str(row.get('Payment Terms Name', ''))
        
        # Tenta encontrar no formato '3x', '12X', etc.
        match = re.search(r'\b(\d{1,2})\s*[xX]\b', payment_str, re.IGNORECASE)
        if match:
            return int(match.group(1))
            
        # Tenta encontrar em formatos de integração como 'installments: 3'
        match = re.search(r'(?:installments|parcelas)\s*:\s*(\d{1,2})', payment_str, re.IGNORECASE)
        if match:
            return int(match.group(1))
            
        return 1

    def _calculate_payment_method_taxes(self, df: pd.DataFrame, card_taxes: dict, pix_tax: float, boleto_tax_brl: float) -> Dict:
        """
        Calcula as taxas por método de pagamento aplicando as novas regras em Reais e parcelas.
        """
        stats = {
            'cartao': {'count': 0, 'total': 0, 'tax_amount': 0.0, 'tax_rate_display': 'Variável'},
            'pix': {'count': 0, 'total': 0, 'tax_amount': 0.0, 'tax_rate_display': f"{pix_tax}%"},
            'boleto': {'count': 0, 'total': 0, 'tax_amount': 0.0, 'tax_rate_display': f"Fixo R$ {boleto_tax_brl:.2f}"},
            'outro': {'count': 0, 'total': 0, 'tax_amount': 0.0, 'tax_rate_display': '0%'}
        }
        
        if 'Financial Status' not in df.columns:
            return stats
            
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        for _, row in unique_paid.iterrows():
            payment_method = self._identify_payment_method(row)
            order_total = row.get('Total', 0)
            
            stats[payment_method]['count'] += 1
            stats[payment_method]['total'] += order_total
            
            if payment_method == 'cartao':
                installments = self._identify_installments(row)
                tax_pct = card_taxes.get(installments, card_taxes.get(1, 4.99))
                stats['cartao']['tax_amount'] += order_total * (tax_pct / 100)
            elif payment_method == 'pix':
                stats['pix']['tax_amount'] += order_total * (pix_tax / 100)
            elif payment_method == 'boleto':
                stats['boleto']['tax_amount'] += boleto_tax_brl
                
        return stats

    def _calculate_category_stats(
        self,
        df: pd.DataFrame,
        costs_map: Dict[str, float],
        default_cost: float,
        product_costs_map: Optional[Dict[str, float]] = None
    ) -> Dict[str, Dict]:
        """Calcula estatísticas por categoria."""
        stats = {}
        product_costs_map = product_costs_map or {}
        
        for _, row in df.iterrows():
            product_name = str(row['Lineitem name'])
            quantity = row['Lineitem quantity']
            price = row['Lineitem price']
            
            category = self._identify_category(product_name)
            unit_cost = product_costs_map.get(category, costs_map.get(category, default_cost))
            
            if category not in stats:
                stats[category] = {"qty": 0, "value": 0, "cost": 0}
            
            stats[category]["qty"] += quantity
            stats[category]["value"] += (price * quantity)
            stats[category]["cost"] += (unit_cost * quantity)
        
        return stats

    def _analyze_repeat_customers(self, df: pd.DataFrame) -> Dict:
        """Analisa clientes que repetiram compra."""
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        if 'Email' not in df.columns:
            return {
                "total_unique_customers": len(unique_paid),
                "repeat_customers": 0,
                "repeat_percentage": 0.0,
                "top_customers": []
            }
        
        email_counts = unique_paid['Email'].value_counts()
        repeat_customers = email_counts[email_counts > 1]
        
        top_customers = []
        for email, count in email_counts.head(10).items():
            customer_orders = unique_paid[unique_paid['Email'] == email]
            total_spent = customer_orders['Total'].sum()
            top_customers.append({
                'email': email,
                'orders': count,
                'total_spent': total_spent,
                'avg_order': total_spent / count
            })
        
        return {
            "total_unique_customers": len(email_counts),
            "repeat_customers": len(repeat_customers),
            "repeat_percentage": (len(repeat_customers) / len(email_counts) * 100) if len(email_counts) > 0 else 0.0,
            "top_customers": top_customers
        }

    def _analyze_timeline(self, df: pd.DataFrame) -> Dict:
        """Analisa vendas por período."""
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        if 'Created at' not in df.columns:
            return {
                "daily_sales": {}, "weekly_sales": {}, "monthly_sales": {},
                "best_day": None, "best_week": None
            }
        
        df_copy = df.copy()
        df_copy['Created at'] = pd.to_datetime(df_copy['Created at'], errors='coerce')
        unique_paid = df_copy[df_copy['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        daily = unique_paid.groupby(unique_paid['Created at'].dt.date)['Total'].agg(['sum', 'count'])
        daily_sales = {str(date): {'value': value, 'count': count} for date, value, count in zip(daily.index, daily['sum'], daily['count'])}
        
        weekly = unique_paid.groupby(unique_paid['Created at'].dt.isocalendar().week)['Total'].agg(['sum', 'count'])
        weekly_sales = {f"Semana {week}": {'value': value, 'count': count} for week, value, count in zip(weekly.index, weekly['sum'], weekly['count'])}
        
        monthly = unique_paid.groupby(unique_paid['Created at'].dt.to_period('M'))['Total'].agg(['sum', 'count'])
        monthly_sales = {str(month): {'value': value, 'count': count} for month, value, count in zip(monthly.index, monthly['sum'], monthly['count'])}
        
        best_day = max(daily_sales.items(), key=lambda x: x[1]['value']) if daily_sales else None
        best_week = max(weekly_sales.items(), key=lambda x: x[1]['value']) if weekly_sales else None
        
        return {
            "daily_sales": daily_sales, "weekly_sales": weekly_sales, "monthly_sales": monthly_sales,
            "best_day": best_day, "best_week": best_week
        }

    def _analyze_geographic(self, df: pd.DataFrame) -> Dict:
        """Analisa vendas por localização geográfica."""
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        geo_data = {
            "by_state": {}, "by_city": {}, "top_states": [], "top_cities": []
        }
        
        if 'Shipping Province' not in df.columns and 'Billing Province' not in df.columns:
            return geo_data
        
        province_col = 'Shipping Province' if 'Shipping Province' in df.columns else 'Billing Province'
        city_col = 'Shipping City' if 'Shipping City' in df.columns else 'Billing City'
        
        if province_col in unique_paid.columns:
            state_sales = unique_paid.groupby(province_col).agg({
                'Total': ['sum', 'count'], 'Subtotal': 'sum'
            }).reset_index()
            state_sales.columns = ['state', 'total_value', 'order_count', 'subtotal']
            state_sales = state_sales.sort_values('total_value', ascending=False)
            geo_data['by_state'] = state_sales.to_dict('records')
            geo_data['top_states'] = state_sales.head(5).values.tolist()
        
        if city_col in unique_paid.columns:
            city_sales = unique_paid.groupby(city_col).agg({
                'Total': ['sum', 'count'], 'Subtotal': 'sum'
            }).reset_index()
            city_sales.columns = ['city', 'total_value', 'order_count', 'subtotal']
            city_sales = city_sales.sort_values('total_value', ascending=False)
            geo_data['by_city'] = city_sales.to_dict('records')
            geo_data['top_cities'] = city_sales.head(5).values.tolist()
        
        return geo_data

    def _analyze_fulfillment(self, df: pd.DataFrame) -> Dict:
        """Analisa status de fulfillment."""
        if 'Fulfillment Status' not in df.columns:
            return {
                "total_orders": 0, "fulfilled": 0, "unfulfilled": 0,
                "partial": 0, "cancelled": 0, "fulfillment_rate": 0.0, "pending_fulfillment": []
            }
        
        unique_orders = df.drop_duplicates(subset=['Name'])
        fulfillment_counts = unique_orders['Fulfillment Status'].value_counts()
        
        total = len(unique_orders)
        fulfilled = fulfillment_counts.get('fulfilled', 0)
        unfulfilled = fulfillment_counts.get('unfulfilled', 0)
        partial = fulfillment_counts.get('partially fulfilled', 0)
        cancelled = fulfillment_counts.get('cancelled', 0)
        
        fulfillment_rate = (fulfilled / total * 100) if total > 0 else 0.0
        
        pending = df[(df['Fulfillment Status'] == 'unfulfilled') & 
                    (df['Financial Status'].str.lower() == 'paid')].drop_duplicates(subset=['Name'])
        
        pending_fulfillment = []
        if 'Created at' in df.columns:
            pending['Created at'] = pd.to_datetime(pending['Created at'], errors='coerce')
            pending_fulfillment = pending[['Name', 'Created at', 'Total']].sort_values('Created at').head(10).values.tolist()
        
        return {
            "total_orders": total, "fulfilled": fulfilled, "unfulfilled": unfulfilled,
            "partial": partial, "cancelled": cancelled, "fulfillment_rate": fulfillment_rate,
            "pending_fulfillment": pending_fulfillment
        }

    def _analyze_discount_codes(self, df: pd.DataFrame) -> Dict:
        """Analisa uso e efetividade de cupons."""
        if 'Discount Code' not in df.columns:
            return {
                "total_discounts": 0, "total_discount_value": 0.0, "discount_codes": [], "top_codes": []
            }
        
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        with_discount = unique_paid[unique_paid['Discount Code'].notna() & (unique_paid['Discount Code'] != '')]
        
        discount_stats = with_discount.groupby('Discount Code').agg({
            'Discount Amount': ['sum', 'count'], 'Total': ['sum', 'mean']
        }).reset_index()
        
        discount_stats.columns = ['code', 'total_discount', 'usage_count', 'total_value', 'avg_order_value']
        discount_stats = discount_stats.sort_values('total_discount', ascending=False)
        
        total_discounts = discount_stats['usage_count'].sum()
        total_discount_value = discount_stats['total_discount'].sum()
        
        return {
            "total_discounts": int(total_discounts),
            "total_discount_value": total_discount_value,
            "discount_codes": discount_stats.to_dict('records'),
            "top_codes": discount_stats.head(5).to_dict('records')
        }

    def process_data(
        self,
        df: pd.DataFrame,
        store_name: str,
        costs_map: Dict[str, float],
        product_costs_map: Optional[Dict[str, float]],
        default_cost: float,
        card_taxes: dict,
        pix_tax: float,
        boleto_tax_brl: float,
        gateway_antifraude_brl: float,
        platform_tax: float,
        ads_cost: float,
        traffic_manager_cost: float
    ) -> Dict:
        """
        Processa os dados e retorna análise.
        """
        df = df.copy()
        
        numeric_cols = ['Subtotal', 'Shipping', 'Total', 'Lineitem quantity', 'Lineitem price']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        paid = df[df['Financial Status'].str.lower() == 'paid']
        cancelled = df[df['Financial Status'].str.lower() == 'cancelled']
        pending = df[df['Financial Status'].str.lower() == 'pending']

        unique_paid = paid.drop_duplicates(subset=['Name'])
        total_items = unique_paid['Subtotal'].sum()
        total_shipping = unique_paid['Shipping'].sum()
        total_received = unique_paid['Total'].sum()

        payment_stats = self._calculate_payment_method_taxes(df, card_taxes, pix_tax, boleto_tax_brl)
        total_payment_taxes = sum(stats['tax_amount'] for stats in payment_stats.values())

        stats = self._calculate_category_stats(paid, costs_map, default_cost, product_costs_map)
        total_prod_cost = sum(cat['cost'] for cat in stats.values())
        
        total_platform_tax = total_received * platform_tax
        total_gateway_antifraude = len(unique_paid) * gateway_antifraude_brl
        
        total_all_taxes = total_payment_taxes + total_platform_tax + total_gateway_antifraude
        total_marketing_costs = ads_cost + traffic_manager_cost
        
        net_profit = total_received - total_all_taxes - total_prod_cost - total_marketing_costs

        free_shipping_orders = unique_paid[unique_paid['Shipping'] == 0]
        free_shipping_count = len(free_shipping_orders)
        free_shipping_value = free_shipping_orders['Total'].sum()
        
        courier_col = None
        transpose_cols = ['Shipping Method', 'Shipping Name', 'Fulfillment Method', 'Carrier', 'Transportadora']
        
        for col in transpose_cols:
            if col in df.columns:
                courier_col = col
                break
        
        correios_pac = 0
        correios_sedex = 0
        transportadoras = 0
        other_couriers = 0
        
        if courier_col is not None:
            for _, order in unique_paid.iterrows():
                shipping_method = str(order[courier_col]).strip().lower() if pd.notna(order[courier_col]) else ""
                if 'pac' in shipping_method: correios_pac += 1
                elif 'sedex' in shipping_method: correios_sedex += 1
                elif shipping_method and shipping_method != 'nan' and order['Shipping'] > 0: transportadoras += 1

        repeat_customers = self._analyze_repeat_customers(df)
        timeline = self._analyze_timeline(df)
        geographic = self._analyze_geographic(df)
        fulfillment = self._analyze_fulfillment(df)
        discounts = self._analyze_discount_codes(df)

        sales_period = {"start_date": None, "end_date": None}
        if 'Created at' in df.columns:
            df_dates = df.copy()
            df_dates['Created at'] = pd.to_datetime(df_dates['Created at'], errors='coerce')
            paid_dates = df_dates[df_dates['Financial Status'].str.lower() == 'paid']['Created at'].dropna()
            if len(paid_dates) > 0:
                sales_period["start_date"] = paid_dates.min().strftime('%d/%m/%Y')
                sales_period["end_date"] = paid_dates.max().strftime('%d/%m/%Y')

        return {
            "store_name": store_name,
            "paid_count": len(unique_paid),
            "cancelled_count": len(cancelled.drop_duplicates(subset=['Name'])),
            "pending_count": len(pending.drop_duplicates(subset=['Name'])),
            "total_items": total_items,
            "total_shipping": total_shipping,
            "total_received": total_received,
            "stats": stats,
            "payment_stats": payment_stats,
            "total_taxes": total_all_taxes,
            "payment_method_taxes": total_payment_taxes,
            "platform_tax_amount": total_platform_tax,
            "gateway_antifraude_amount": total_gateway_antifraude,
            "total_prod_cost": total_prod_cost,
            "ads_cost": ads_cost,
            "traffic_manager_cost": traffic_manager_cost,
            "total_marketing_costs": total_marketing_costs,
            "net_profit": net_profit,
            "analysis_date": datetime.now(),
            "free_shipping_count": free_shipping_count,
            "free_shipping_value": free_shipping_value,
            "correios_pac": correios_pac,
            "correios_sedex": correios_sedex,
            "transportadoras": transportadoras,
            "other_couriers": other_couriers,
            "repeat_customers": repeat_customers,
            "timeline": timeline,
            "geographic": geographic,
            "fulfillment": fulfillment,
            "discounts": discounts,
            "sales_period": sales_period
        }

    def _calculate_roi_by_coupon(self, df: pd.DataFrame) -> Dict:
        """Calcula ROI para cada cupom."""
        if 'Discount Code' not in df.columns:
            return {}
        
        unique_paid = df[df['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        without_discount = unique_paid[unique_paid['Discount Code'].isna() | (unique_paid['Discount Code'] == '')]
        avg_ticket_without = without_discount['Total'].mean() if len(without_discount) > 0 else 0
        
        roi_data = {}
        with_discount = unique_paid[unique_paid['Discount Code'].notna() & (unique_paid['Discount Code'] != '')]
        
        for code in with_discount['Discount Code'].unique():
            code_orders = with_discount[with_discount['Discount Code'] == code]
            total_discount = code_orders['Discount Amount'].sum()
            total_revenue = code_orders['Total'].sum()
            order_count = len(code_orders)
            avg_ticket_with = code_orders['Total'].mean()
            
            roi = ((total_revenue - total_discount) / total_discount * 100) if total_discount > 0 else 0
            ticket_increase = ((avg_ticket_with - avg_ticket_without) / avg_ticket_without * 100) if avg_ticket_without > 0 else 0
            
            roi_data[str(code)] = {
                'code': code, 'orders': order_count, 'total_discount': total_discount,
                'total_revenue': total_revenue, 'roi': roi, 'avg_ticket': avg_ticket_with,
                'ticket_increase': ticket_increase
            }
        return roi_data

    def _compare_periods(self, df: pd.DataFrame, date_from1: str, date_to1: str, date_from2: str, date_to2: str) -> Dict:
        """Compara dois períodos."""
        df_copy = df.copy()
        df_copy['Created at'] = pd.to_datetime(df_copy['Created at'], errors='coerce')
        unique_paid = df_copy[df_copy['Financial Status'].str.lower() == 'paid'].drop_duplicates(subset=['Name'])
        
        p1 = unique_paid[(unique_paid['Created at'] >= date_from1) & (unique_paid['Created at'] <= date_to1)]
        p2 = unique_paid[(unique_paid['Created at'] >= date_from2) & (unique_paid['Created at'] <= date_to2)]
        
        return {
            'period1': {
                'orders': len(p1), 'revenue': p1['Total'].sum(), 'avg_ticket': p1['Total'].mean(), 'items': p1['Subtotal'].sum()
            },
            'period2': {
                'orders': len(p2), 'revenue': p2['Total'].sum(), 'avg_ticket': p2['Total'].mean(), 'items': p2['Subtotal'].sum()
            },
            'growth': {
                'orders_pct': ((len(p2) - len(p1)) / len(p1) * 100) if len(p1) > 0 else 0,
                'revenue_pct': ((p2['Total'].sum() - p1['Total'].sum()) / p1['Total'].sum() * 100) if p1['Total'].sum() > 0 else 0,
                'ticket_pct': ((p2['Total'].mean() - p1['Total'].mean()) / p1['Total'].mean() * 100) if p1['Total'].mean() > 0 else 0
            }
        }

    def generate_excel_report(self, analysis_data: Dict) -> bytes:
        """Gera relatório em Excel."""
        output = io.BytesIO()
        wb = Workbook()
        
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        
        ws = wb.active
        ws.title = "Resumo"
        ws['A1'] = "RESUMO EXECUTIVO"
        ws['A1'].font = title_font
        
        if analysis_data.get('sales_period') and analysis_data['sales_period'].get('start_date'):
            ws['A2'] = f"Período: {analysis_data['sales_period']['start_date']} até {analysis_data['sales_period']['end_date']}"
        
        row = 4
        metrics = [
            ("Pedidos Pagos", analysis_data['paid_count']),
            ("Pedidos Cancelados", analysis_data['cancelled_count']),
            ("Pedidos Pendentes", analysis_data['pending_count']),
            ("Subtotal", analysis_data['total_items']),
            ("Frete Total", analysis_data['total_shipping']),
            ("Total Recebido", analysis_data['total_received']),
            ("Taxa Gateway & Antifraude (R$)", analysis_data['gateway_antifraude_amount']),
            ("Frete Grátis", analysis_data['free_shipping_count']),
            ("Lucro Líquido", analysis_data['net_profit']),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value if isinstance(value, (int, float)) and value < 1000 else value
            row += 1
        
        ws = wb.create_sheet("Métodos de Pagamento")
        ws['A1'] = "MÉTODO"
        ws['B1'] = "QUANTIDADE"
        ws['C1'] = "TOTAL (R$)"
        ws['D1'] = "TAXA"
        ws['E1'] = "TAXA COBRADA (R$)"
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        payment_method_names = {'cartao': 'Cartão de Crédito', 'pix': 'Pix', 'boleto': 'Boleto', 'outro': 'Outro'}
        
        for method, data in analysis_data['payment_stats'].items():
            ws[f'A{row}'] = payment_method_names.get(method, method)
            ws[f'B{row}'] = int(data['count'])
            ws[f'C{row}'] = data['total']
            ws[f'D{row}'] = data['tax_rate_display']
            ws[f'E{row}'] = data.get('tax_amount', 0)
            row += 1
        
        ws = wb.create_sheet("Categorias")
        ws['A1'] = "CATEGORIA"
        ws['B1'] = "QUANTIDADE"
        ws['C1'] = "VENDA (R$)"
        ws['D1'] = "CUSTO (R$)"
        ws['E1'] = "MARGEM (R$)"
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for cat, data in sorted(analysis_data['stats'].items(), key=lambda x: x[1]['value'], reverse=True):
            ws[f'A{row}'] = cat
            ws[f'B{row}'] = int(data['qty'])
            ws[f'C{row}'] = data['value']
            ws[f'D{row}'] = data['cost']
            ws[f'E{row}'] = data['value'] - data['cost']
            row += 1
        
        ws = wb.create_sheet("Top Clientes")
        ws['A1'] = "EMAIL"
        ws['B1'] = "PEDIDOS"
        ws['C1'] = "GASTO TOTAL (R$)"
        ws['D1'] = "TICKET MÉDIO (R$)"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        row = 2
        for customer in analysis_data['repeat_customers'].get('top_customers', []):
            ws[f'A{row}'] = customer['email']
            ws[f'B{row}'] = customer['orders']
            ws[f'C{row}'] = customer['total_spent']
            ws[f'D{row}'] = customer['avg_order']
            row += 1
        
        ws = wb.create_sheet("Geográfica")
        ws['A1'] = "ESTADO"
        ws['B1'] = "TOTAL (R$)"
        ws['C1'] = "PEDIDOS"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        row = 2
        for state in analysis_data['geographic'].get('top_states', []):
            ws[f'A{row}'] = state[0]
            ws[f'B{row}'] = state[1]
            ws[f'C{row}'] = int(state[2])
            row += 1
        
        if analysis_data['discounts'].get('top_codes'):
            ws = wb.create_sheet("Cupons")
            ws['A1'] = "CUPOM"
            ws['B1'] = "USOS"
            ws['C1'] = "DESC. TOTAL (R$)"
            ws['D1'] = "TICKET MÉD (R$)"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            row = 2
            for code in analysis_data['discounts']['top_codes']:
                ws[f'A{row}'] = code['code']
                ws[f'B{row}'] = int(code['usage_count'])
                ws[f'C{row}'] = code['total_discount']
                ws[f'D{row}'] = code['avg_order_value']
                row += 1
        
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf(analysis_data: Dict) -> bytes:
        """Gera PDF com a análise."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'], fontSize=16,
            textColor=colors.HexColor('#1f77b4'), spaceAfter=30, alignment=1
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'], fontSize=12,
            textColor=colors.HexColor('#1f77b4'), spaceAfter=12, spaceBefore=12
        )
        
        title = f"Relatório de Vendas - {analysis_data['store_name'].upper()}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        date_text = f"Data: {analysis_data['analysis_date'].strftime('%d/%m/%Y %H:%M')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        
        if analysis_data.get('sales_period') and analysis_data['sales_period'].get('start_date'):
            period_text = f"Período das Vendas: {analysis_data['sales_period']['start_date']} até {analysis_data['sales_period']['end_date']}"
            elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("VOLUME DE PEDIDOS", heading_style))
        data = [
            ['Pagos', str(analysis_data['paid_count'])],
            ['Cancelados', str(analysis_data['cancelled_count'])],
            ['Pendentes', str(analysis_data['pending_count'])]
        ]
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("VALORES TOTAIS", heading_style))
        data = [
            ['Subtotal (Produtos)', f"R$ {analysis_data['total_items']:,.2f}"],
            ['Frete Total', f"R$ {analysis_data['total_shipping']:,.2f}"],
            ['Total Recebido', f"R$ {analysis_data['total_received']:,.2f}"]
        ]
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("ANÁLISE DE FRETE E TRANSPORTADORAS", heading_style))
        data = [
            ['Frete Grátis (Qty)', str(analysis_data.get('free_shipping_count', 0))],
            ['Frete Grátis (Valor)', f"R$ {analysis_data.get('free_shipping_value', 0):,.2f}"],
            ['Pedidos - Correios PAC', str(analysis_data.get('correios_pac', 0))],
            ['Pedidos - Correios SEDEX', str(analysis_data.get('correios_sedex', 0))],
            ['Pedidos - Transportadoras', str(analysis_data.get('transportadoras', 0))],
            ['Pedidos - Outros', str(analysis_data.get('other_couriers', 0))]
        ]
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightyellow),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("ANÁLISE DE MÉTODOS DE PAGAMENTO", heading_style))
        payment_method_names = {'cartao': 'Cartão de Crédito', 'pix': 'Pix', 'boleto': 'Boleto', 'outro': 'Outro'}
        data = [['Método', 'Pedidos', 'Total (R$)', 'Taxa', 'Imposto (R$)']]
        for method, stats in analysis_data['payment_stats'].items():
            data.append([
                payment_method_names.get(method, method),
                str(int(stats['count'])),
                f"{stats['total']:,.2f}",
                stats['tax_rate_display'],
                f"{stats.get('tax_amount', 0):,.2f}"
            ])
        table = Table(data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("DETALHAMENTO POR CATEGORIA", heading_style))
        sorted_stats = sorted(analysis_data['stats'].items(), key=lambda x: x[1]['value'], reverse=True)
        data = [['Categoria', 'Qtd', 'Venda (R$)', 'Custo (R$)']]
        for category, cat_data in sorted_stats:
            data.append([category, str(int(cat_data['qty'])), f"{cat_data['value']:,.2f}", f"{cat_data['cost']:,.2f}"])
        table = Table(data, colWidths=[2*inch, 1*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("RESUMO FINANCEIRO", heading_style))
        data = [
            ['(-) Taxas por Método de Pagamento', f"R$ {analysis_data['payment_method_taxes']:,.2f}"],
            ['(-) Taxa Plataforma', f"R$ {analysis_data['platform_tax_amount']:,.2f}"],
            ['(-) Gateway & Antifraude', f"R$ {analysis_data['gateway_antifraude_amount']:,.2f}"],
            ['(-) Custo Produção', f"R$ {analysis_data['total_prod_cost']:,.2f}"],
            ['(-) Gasto ADS', f"R$ {analysis_data['ads_cost']:,.2f}"],
            ['(-) Gestor de Tráfego', f"R$ {analysis_data['traffic_manager_cost']:,.2f}"],
            ['(=) LUCRO LÍQUIDO', f"R$ {analysis_data['net_profit']:,.2f}"]
        ]
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -2), colors.lightcyan),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.darkgreen),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


def save_analysis_history(analysis_data: Dict) -> None:
    """Salva análise no histórico."""
    history = []
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except:
            history = []
    
    analysis_entry = {
        'timestamp': datetime.now().isoformat(),
        'store': analysis_data['store_name'],
        'paid_orders': analysis_data['paid_count'],
        'revenue': float(analysis_data['total_received']),
        'profit': float(analysis_data['net_profit']),
        'shipping': float(analysis_data['total_shipping'])
    }
    history.append(analysis_entry)
    if len(history) > 100: history = history[-100:]
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)


def load_analysis_history() -> list:
    """Carrega histórico de análises."""
    if not os.path.exists(HISTORY_FILE): return []
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except: return []


def create_sales_chart(timeline_data: Dict) -> go.Figure:
    if not timeline_data['daily_sales']: return None
    dates, values = [], []
    for date, data in sorted(timeline_data['daily_sales'].items()):
        dates.append(date); values.append(data['value'])
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=dates, y=values, mode='lines+markers', name='Vendas Diárias', line=dict(color='#1f77b4', width=2), marker=dict(size=6)))
    fig.update_layout(title="📈 Vendas Diárias", xaxis_title="Data", yaxis_title="Valor (R$)", hovermode='x unified', template='plotly_white')
    return fig


def create_category_chart(stats: Dict) -> go.Figure:
    categories, values = [], []
    for cat, data in sorted(stats.items(), key=lambda x: x[1]['value'], reverse=True):
        categories.append(cat); values.append(data['value'])
    fig = px.bar(x=categories, y=values, labels={'x': 'Categoria', 'y': 'Vendas (R$)'}, title="📊 Vendas por Categoria", color=values, color_continuous_scale='Blues')
    fig.update_layout(template='plotly_white')
    return fig


def create_geographic_chart(geo_data: Dict) -> go.Figure:
    if not geo_data['top_states']: return None
    states, values = [], []
    for state in geo_data['top_states']:
        states.append(state[0]); values.append(state[1])
    fig = px.pie(values=values, names=states, title="🗺️ Distribuição de Vendas por Estado")
    fig.update_layout(template='plotly_white')
    return fig


def create_fulfillment_chart(fulfillment: Dict) -> go.Figure:
    labels = ['Entregues', 'Não Entregues']
    sizes = [fulfillment['fulfilled'], fulfillment['unfulfilled']]
    colors = ['#4CAF50', '#FF9800']
    fig = go.Figure(data=[go.Pie(labels=labels, values=sizes, marker=dict(colors=colors))])
    fig.update_layout(title="📦 Status de Fulfillment", template='plotly_white')
    return fig


def create_coupon_chart(discounts: Dict) -> go.Figure:
    if not discounts['top_codes']: return None
    codes, uses = [], []
    for code in discounts['top_codes']:
        codes.append(code['code']); uses.append(code['usage_count'])
    fig = px.bar(x=codes, y=uses, labels={'x': 'Cupom', 'y': 'Utilizações'}, title="💳 Top Cupons", color=uses, color_continuous_scale='Greens')
    fig.update_layout(template='plotly_white')
    return fig


def main():
    """Função principal da aplicação web."""
    st.title("📊 Analisador de Vendas Multi-Loja")
    
    if 'custom_categories' not in st.session_state:
        st.session_state.custom_categories = DEFAULT_CATEGORIAS_CONFIG.copy()
    else:
        for category_name, keywords in DEFAULT_CATEGORIAS_CONFIG.items():
            if category_name not in st.session_state.custom_categories:
                st.session_state.custom_categories[category_name] = keywords
    
    analyzer = VendasAnalyzerWeb(st.session_state.custom_categories)
    
    tab1, tab2, tab3, tab4 = st.tabs(["📈 Análise Básica", "🔍 Análises Avançadas", "⚙️ Categorias", "ℹ️ Ajuda"])
    
    with tab1:
        st.markdown("### 1️⃣ Informações Básicas")
        col1, col2 = st.columns([2, 1])
        with col1:
            store_name = st.text_input("Nome da Loja", value=DEFAULT_STORE_NAME, help="Nome da loja para o relatório")
        
        st.markdown("### 2️⃣ Seleção de Arquivo")
        uploaded_file = st.file_uploader("Selecione seu arquivo CSV de vendas", type=['csv'], help="Arquivo exportado da Shopify ou plataforma de vendas")
        
        if uploaded_file is not None:
            try:
                for encoding in ['utf-8', 'latin-1', 'iso-8859-1']:
                    try:
                        df = pd.read_csv(uploaded_file, encoding=encoding)
                        st.success(f"✅ Arquivo carregado com sucesso ({len(df)} linhas)")
                        break
                    except UnicodeDecodeError:
                        continue
                
                st.markdown("### 3️⃣ Custos de Produção (R$)")
                col1, col2, col3 = st.columns(3)
                costs = {}
                categories = analyzer.get_categories_list()
                
                with col1:
                    for cat in categories[:2]:
                        default = DEFAULT_OVERSIZED_COST if cat == "Oversized" else 0.0
                        costs[cat] = st.number_input(cat, value=float(default), min_value=0.0, step=0.01, key=f"cost_{cat}")
                
                with col2:
                    for cat in categories[2:4]:
                        costs[cat] = st.number_input(cat, value=0.0, min_value=0.0, step=0.01, key=f"cost_{cat}")
                
                with col3:
                    for cat in categories[4:]:
                        costs[cat] = st.number_input(cat, value=0.0, min_value=0.0, step=0.01, key=f"cost_{cat}")
                    default_cost = st.number_input("Outros (padrão)", value=0.0, min_value=0.0, step=0.01)

                product_costs = {}
                detected_categories = []
                if 'Lineitem name' in df.columns:
                    detected_categories = sorted({
                        analyzer._identify_category(name)
                        for name in df['Lineitem name'].dropna().astype(str)
                        if str(name).strip() and analyzer._identify_category(name) != "Outros"
                    })

                st.markdown("### 3.1️⃣ Custos por Tipo de Produto (Automático do CSV)")
                st.caption(f"Tipos detectados no upload: {len(detected_categories)}")
                if detected_categories:
                    with st.expander("🧾 Definir custo por tipo detectado", expanded=True):
                        for idx, category_name in enumerate(detected_categories):
                            suggested_cost = costs.get(category_name, default_cost)
                            label = f"{category_name}"
                            product_costs[category_name] = st.number_input(
                                label,
                                value=float(suggested_cost),
                                min_value=0.0,
                                step=0.01,
                                key=f"product_cost_{idx}"
                            )
                else:
                    st.info("Nenhum tipo conhecido foi identificado no CSV para custo automático.")
                
                st.markdown("### 4️⃣ Taxas e Custos")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown("#### Taxas de Pagamento")
                    with st.expander("💳 Taxas de Cartão (Por Parcelas)"):
                        card_taxes = {}
                        for i in range(1, 13):
                            default_val = 4.99 if i == 1 else 4.99 + (i * 0.5) 
                            card_taxes[i] = st.number_input(f"Taxa {i}x (%)", value=default_val, min_value=0.0, step=0.01, key=f"card_tax_{i}")
                    
                    pix_tax = st.number_input("Taxa Pix (%)", value=float(DEFAULT_PIX_TAX), min_value=0.0, step=0.01)
                    boleto_tax = st.number_input("Taxa Boleto (R$ Fixo por venda)", value=float(DEFAULT_BOLETO_TAX_BRL), min_value=0.0, step=0.01)

                with col2:
                    st.markdown("#### Taxas Operacionais")
                    platform_tax = st.number_input("Taxa da Plataforma (%)", value=0.0, min_value=0.0, step=0.01)
                    gateway_antifraude_tax = st.number_input("Gateway e Antifraude (R$ Fixo por venda)", value=float(DEFAULT_GATEWAY_TAX_BRL), min_value=0.0, step=0.01)

                with col3:
                    st.markdown("#### Custos de Marketing (R$)")
                    ads_cost = st.number_input("Gasto Total com ADS", value=0.0, min_value=0.0, step=0.01)
                    traffic_manager_cost = st.number_input("Custo Gestor de Tráfego", value=0.0, min_value=0.0, step=0.01)
                
                if st.button("🚀 GERAR ANÁLISE COMPLETA", use_container_width=True):
                    try:
                        analysis_data = analyzer.process_data(
                            df=df,
                            store_name=store_name,
                            costs_map=costs,
                            product_costs_map=product_costs,
                            default_cost=default_cost,
                            card_taxes=card_taxes,
                            pix_tax=pix_tax,
                            boleto_tax_brl=boleto_tax,
                            gateway_antifraude_brl=gateway_antifraude_tax,
                            platform_tax=platform_tax/100,
                            ads_cost=ads_cost,
                            traffic_manager_cost=traffic_manager_cost
                        )
                        
                        st.session_state.analysis_data = analysis_data
                        st.session_state.df_analysis = df
                        st.session_state.show_results = True
                        save_analysis_history(analysis_data)
                        st.success("✅ Análise realizada com sucesso!")
                        
                    except Exception as e:
                        st.error(f"❌ Erro na análise: {str(e)}")
                
                if st.session_state.get('show_results'):
                    analysis = st.session_state.analysis_data
                    
                    st.markdown("---")
                    st.markdown("## 📊 Resultado da Análise")
                    
                    if analysis.get('sales_period') and analysis['sales_period'].get('start_date'):
                        period_text = f"📅 **Período das Vendas:** {analysis['sales_period']['start_date']} até {analysis['sales_period']['end_date']}"
                        st.info(period_text)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1: st.metric("Pedidos Pagos", analysis['paid_count'])
                    with col2: st.metric("Pedidos Cancelados", analysis['cancelled_count'])
                    with col3: st.metric("Pedidos Pendentes", analysis['pending_count'])
                    with col4: st.metric("Lucro Líquido", f"R$ {analysis['net_profit']:,.2f}", delta=f"{(analysis['net_profit']/analysis['total_received']*100):.1f}%" if analysis['total_received'] > 0 else "0%")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1: st.metric("Subtotal", f"R$ {analysis['total_items']:,.2f}")
                    with col2: st.metric("Frete", f"R$ {analysis['total_shipping']:,.2f}")
                    with col3: st.metric("Total Recebido", f"R$ {analysis['total_received']:,.2f}")
                    
                    st.markdown("### 💳 Análise de Métodos de Pagamento")
                    payment_method_names = {'cartao': 'Cartão de Crédito', 'pix': 'Pix', 'boleto': 'Boleto', 'outro': 'Outro'}
                    
                    payment_data = []
                    for method, stats in analysis['payment_stats'].items():
                        if stats['count'] > 0:
                            payment_data.append({
                                'Método': payment_method_names.get(method, method),
                                'Pedidos': int(stats['count']),
                                'Total (R$)': f"R$ {stats['total']:,.2f}",
                                'Taxa Configurada': stats['tax_rate_display'],
                                'Imposto (R$)': f"R$ {stats.get('tax_amount', 0):,.2f}"
                            })
                    
                    if payment_data:
                        st.dataframe(pd.DataFrame(payment_data), use_container_width=True, hide_index=True)
                    
                    st.markdown("### 📦 Análise de Frete e Transportadoras")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1: st.metric("Frete Grátis", f"{analysis['free_shipping_count']} pedidos", f"R$ {analysis['free_shipping_value']:,.2f}")
                    with col2: st.metric("Correios PAC", f"{analysis['correios_pac']} pedidos")
                    with col3: st.metric("Correios SEDEX", f"{analysis['correios_sedex']} pedidos")
                    with col4: st.metric("Transportadoras", f"{analysis['transportadoras']} pedidos")
                    if analysis['other_couriers'] > 0: st.info(f"📌 {analysis['other_couriers']} pedido(s) com outras transportadoras")
                    
                    st.markdown("### Resumo Financeiro")
                    financial_data = {
                        'Taxas Pagamento': analysis['payment_method_taxes'],
                        'Taxa Plataforma': analysis['platform_tax_amount'],
                        'Gateway / Antifraude': analysis['gateway_antifraude_amount'],
                        'Custo Produção': analysis['total_prod_cost'],
                        'Gasto ADS': analysis['ads_cost'],
                        'Gestor de Tráfego': analysis['traffic_manager_cost'],
                    }
                    col1, col2 = st.columns([1, 1])
                    with col1:
                        st.write("**Despesas:**")
                        for item, value in financial_data.items():
                            st.write(f"- {item}: R$ {value:,.2f}")
                    with col2:
                        st.write("**Resultado:**")
                        color = '🟢' if analysis['net_profit'] > 0 else '🔴'
                        st.write(f"{color} **Lucro Líquido: R$ {analysis['net_profit']:,.2f}**")
                    
                    st.markdown("### Detalhamento por Categoria")
                    category_data = []
                    for cat, data in sorted(analysis['stats'].items(), key=lambda x: x[1]['value'], reverse=True):
                        category_data.append({
                            'Categoria': cat, 'Quantidade': int(data['qty']),
                            'Venda': f"R$ {data['value']:,.2f}", 'Custo': f"R$ {data['cost']:,.2f}",
                            'Margem': f"R$ {data['value'] - data['cost']:,.2f}"
                        })
                    st.dataframe(pd.DataFrame(category_data), use_container_width=True, hide_index=True)
                    
                    st.markdown("---")
                    st.markdown("### 📥 Downloads")
                    pdf_data = analyzer.generate_pdf(analysis)
                    excel_data = analyzer.generate_excel_report(analysis)
                    filename_pdf = f"Relatorio_{store_name}_{analysis['analysis_date'].strftime('%Y%m%d_%H%M')}.pdf"
                    filename_excel = f"Relatorio_{store_name}_{analysis['analysis_date'].strftime('%Y%m%d_%H%M')}.xlsx"
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.download_button("📄 PDF", data=pdf_data, file_name=filename_pdf, mime="application/pdf", use_container_width=True)
                    with col2:
                        st.download_button("📊 Excel", data=excel_data, file_name=filename_excel, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    with col3:
                        if st.button("🔄 Nova Análise", use_container_width=True):
                            st.session_state.show_results = False
                            st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao processar arquivo: {str(e)}")
    
    with tab2:
        st.markdown("## 🔍 Análises Avançadas")
        if not st.session_state.get('show_results') or not st.session_state.get('analysis_data'):
            st.info("⏳ Execute uma análise na aba '📈 Análise Básica' para ver as análises avançadas.")
        else:
            analysis = st.session_state.analysis_data
            subab1, subab2, subab3, subab4, subab5 = st.tabs(["👥 Clientes", "📅 Timeline", "🗺️ Geográfica", "📦 Fulfillment", "💳 Cupons"])
            
            with subab1:
                st.markdown("### 👥 Análise de Clientes")
                repeat = analysis['repeat_customers']
                col1, col2, col3 = st.columns(3)
                with col1: st.metric("Clientes Únicos", repeat['total_unique_customers'])
                with col2: st.metric("Clientes Repeat", repeat['repeat_customers'])
                with col3: st.metric("% Repeat", f"{repeat['repeat_percentage']:.1f}%")
                
                st.markdown("---")
                st.markdown("### Top 10 Clientes")
                if repeat['top_customers']:
                    top_customers_data = []
                    for customer in repeat['top_customers']:
                        top_customers_data.append({
                            'Email': customer['email'], 'Pedidos': customer['orders'],
                            'Gasto Total': f"R$ {customer['total_spent']:,.2f}", 'Ticket Médio': f"R$ {customer['avg_order']:,.2f}"
                        })
                    st.dataframe(pd.DataFrame(top_customers_data), use_container_width=True, hide_index=True)
                else:
                    st.warning("Nenhum dado de cliente disponível")
            
            with subab2:
                st.markdown("### 📅 Análise de Vendas por Período")
                timeline = analysis['timeline']
                if timeline['best_day']:
                    col1, col2 = st.columns(2)
                    with col1:
                        day, day_data = timeline['best_day']
                        st.metric("Melhor Dia", day, f"R$ {day_data['value']:,.2f} ({day_data['count']} pedidos)")
                    with col2:
                        if timeline['best_week']:
                            week, week_data = timeline['best_week']
                            st.metric("Melhor Semana", week, f"R$ {week_data['value']:,.2f} ({week_data['count']} pedidos)")
                
                st.markdown("---")
                if timeline['daily_sales']:
                    st.markdown("#### 📊 Vendas Diárias")
                    daily_data = []
                    for date, data in sorted(timeline['daily_sales'].items(), reverse=True):
                        daily_data.append({'Data': date, 'Valor': f"R$ {data['value']:,.2f}", 'Pedidos': data['count']})
                    st.dataframe(pd.DataFrame(daily_data), use_container_width=True, hide_index=True)
                
                if timeline['monthly_sales']:
                    st.markdown("#### 📈 Vendas Mensais")
                    monthly_data = []
                    for month, data in sorted(timeline['monthly_sales'].items()):
                        monthly_data.append({'Mês': month, 'Valor': f"R$ {data['value']:,.2f}", 'Pedidos': data['count']})
                    st.dataframe(pd.DataFrame(monthly_data), use_container_width=True, hide_index=True)
            
            with subab3:
                st.markdown("### 🗺️ Análise Geográfica")
                geo = analysis['geographic']
                if geo['top_states']:
                    st.markdown("#### 🏆 Top Estados")
                    states_data = []
                    for state in geo['top_states']:
                        states_data.append({'Estado': state[0], 'Total': f"R$ {state[1]:,.2f}", 'Pedidos': int(state[2]), 'Subtotal': f"R$ {state[3]:,.2f}"})
                    st.dataframe(pd.DataFrame(states_data), use_container_width=True, hide_index=True)
                st.markdown("---")
                if geo['top_cities']:
                    st.markdown("#### 🏘️ Top Cidades")
                    cities_data = []
                    for city in geo['top_cities']:
                        cities_data.append({'Cidade': city[0], 'Total': f"R$ {city[1]:,.2f}", 'Pedidos': int(city[2]), 'Subtotal': f"R$ {city[3]:,.2f}"})
                    st.dataframe(pd.DataFrame(cities_data), use_container_width=True, hide_index=True)
                if not geo['top_states'] and not geo['top_cities']: st.warning("Nenhum dado geográfico disponível")
            
            with subab4:
                st.markdown("### 📦 Status de Fulfillment")
                fulfill = analysis['fulfillment']
                col1, col2, col3, col4 = st.columns(4)
                with col1: st.metric("Total de Pedidos", fulfill['total_orders'])
                with col2: st.metric("Entregues", fulfill['fulfilled'])
                with col3: st.metric("Não Entregues", fulfill['unfulfilled'])
                with col4: st.metric("Taxa Entrega", f"{fulfill['fulfillment_rate']:.1f}%")
                st.markdown("---")
                if fulfill['partial'] > 0: st.info(f"⚠️ {fulfill['partial']} pedido(s) parcialmente entregue(s)")
                if fulfill['cancelled'] > 0: st.warning(f"❌ {fulfill['cancelled']} pedido(s) cancelado(s)")
                st.markdown("---")
                if fulfill['pending_fulfillment']:
                    st.markdown("#### ⏳ Pedidos Pendentes de Entrega")
                    pending_data = []
                    for order in fulfill['pending_fulfillment']:
                        pending_data.append({'Pedido': order[0], 'Data': order[1], 'Valor': f"R$ {order[2]:,.2f}"})
                    st.dataframe(pd.DataFrame(pending_data), use_container_width=True, hide_index=True)
            
            with subab5:
                st.markdown("### 💳 Análise de Cupons")
                disc = analysis['discounts']
                col1, col2, col3 = st.columns(3)
                with col1: st.metric("Cupons Utilizados", disc['total_discounts'])
                with col2: st.metric("Desconto Total", f"R$ {disc['total_discount_value']:,.2f}")
                with col3:
                    avg_discount = disc['total_discount_value'] / disc['total_discounts'] if disc['total_discounts'] > 0 else 0
                    st.metric("Desconto Médio", f"R$ {avg_discount:,.2f}")
                st.markdown("---")
                if disc['top_codes']:
                    st.markdown("#### 🏆 Top Cupons")
                    codes_data = []
                    for code in disc['top_codes']:
                        codes_data.append({'Cupom': code['code'], 'Usos': int(code['usage_count']), 'Desconto Total': f"R$ {code['total_discount']:,.2f}", 'Ticket Médio': f"R$ {code['avg_order_value']:,.2f}"})
                    st.dataframe(pd.DataFrame(codes_data), use_container_width=True, hide_index=True)
                if not disc['top_codes']: st.info("Nenhum cupom foi usado neste período")
            
            subab6 = st.expander("📊 Gráficos & Visualizações", expanded=False)
            with subab6:
                st.markdown("### 📈 Visualizações Detalhadas")
                sales_chart = create_sales_chart(analysis['timeline'])
                if sales_chart: st.plotly_chart(sales_chart, use_container_width=True)
                cat_chart = create_category_chart(analysis['stats'])
                if cat_chart: st.plotly_chart(cat_chart, use_container_width=True)
                col1, col2 = st.columns(2)
                with col1:
                    geo_chart = create_geographic_chart(analysis['geographic'])
                    if geo_chart: st.plotly_chart(geo_chart, use_container_width=True)
                with col2:
                    fulfill_chart = create_fulfillment_chart(analysis['fulfillment'])
                    if fulfill_chart: st.plotly_chart(fulfill_chart, use_container_width=True)
                coupon_chart = create_coupon_chart(analysis['discounts'])
                if coupon_chart: st.plotly_chart(coupon_chart, use_container_width=True)
            
            subab7 = st.expander("💰 ROI por Cupom", expanded=False)
            with subab7:
                st.markdown("### 💹 Análise de ROI")
                df_for_roi = st.session_state.get('df_analysis', pd.DataFrame())
                roi_data = analyzer._calculate_roi_by_coupon(df_for_roi)
                if roi_data:
                    roi_list = []
                    for code, data in roi_data.items():
                        roi_list.append({'Cupom': data['code'], 'Usos': data['orders'], 'Receita Gerada': f"R$ {data['total_revenue']:,.2f}", 'Desconto Dado': f"R$ {data['total_discount']:,.2f}", 'ROI (%)': f"{data['roi']:.1f}%", 'Aumento Ticket': f"{data['ticket_increase']:+.1f}%"})
                    st.dataframe(pd.DataFrame(roi_list), use_container_width=True, hide_index=True)
                    st.markdown("---")
                    st.markdown("**Interpretação:**")
                    st.text("ROI = (Receita - Desconto) / Desconto × 100%")
                    st.text("Aumento Ticket = (Ticket com cupom - Ticket sem cupom) / Ticket sem cupom × 100%")
                else:
                    st.info("Nenhum dado de cupom disponível para cálculo de ROI")
            
            subab8 = st.expander("📊 Comparar Períodos", expanded=False)
            with subab8:
                st.markdown("### 📅 Comparação entre Períodos")
                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("Período 1")
                    p1_from = st.date_input("De", key="p1_from")
                    p1_to = st.date_input("Até", key="p1_to")
                with col2:
                    st.subheader("Período 2")
                    p2_from = st.date_input("De", key="p2_from")
                    p2_to = st.date_input("Até", key="p2_to")
                
                if st.button("▶️ Comparar Períodos"):
                    try:
                        if 'df_analysis' in st.session_state:
                            comparison = analyzer._compare_periods(st.session_state.df_analysis, pd.Timestamp(p1_from), pd.Timestamp(p1_to), pd.Timestamp(p2_from), pd.Timestamp(p2_to))
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown(f"**Período 1:** {p1_from} a {p1_to}")
                                st.metric("Pedidos", int(comparison['period1']['orders']))
                                st.metric("Receita", f"R$ {comparison['period1']['revenue']:,.2f}")
                                st.metric("Ticket Médio", f"R$ {comparison['period1']['avg_ticket']:,.2f}")
                            with col2:
                                st.markdown(f"**Período 2:** {p2_from} a {p2_to}")
                                st.metric("Pedidos", int(comparison['period2']['orders']), f"{comparison['growth']['orders_pct']:+.1f}%")
                                st.metric("Receita", f"R$ {comparison['period2']['revenue']:,.2f}", f"{comparison['growth']['revenue_pct']:+.1f}%")
                                st.metric("Ticket Médio", f"R$ {comparison['period2']['avg_ticket']:,.2f}", f"{comparison['growth']['ticket_pct']:+.1f}%")
                    except Exception as e:
                        st.error(f"Erro ao comparar períodos: {str(e)}")
            
            subab9 = st.expander("📚 Histórico de Análises", expanded=False)
            with subab9:
                st.markdown("### 📋 Análises Anteriores")
                history = load_analysis_history()
                if history:
                    history_data = []
                    for entry in reversed(history[-20:]):
                        history_data.append({'Data': entry['timestamp'][:10], 'Hora': entry['timestamp'][11:16], 'Loja': entry['store'], 'Pedidos': entry['paid_orders'], 'Receita': f"R$ {entry['revenue']:,.2f}", 'Lucro': f"R$ {entry['profit']:,.2f}"})
                    st.dataframe(pd.DataFrame(history_data), use_container_width=True, hide_index=True)
                    st.markdown("---")
                    if st.button("🗑️ Limpar Histórico"):
                        if os.path.exists(HISTORY_FILE):
                            os.remove(HISTORY_FILE)
                            st.success("Histórico limpo!")
                            st.rerun()
                else:
                    st.info("Nenhuma análise anterior encontrada")
                
                if st.button("💾 Salvar Análise no Histórico"):
                    save_analysis_history(analysis)
                    st.success("✅ Análise salva no histórico!")
    
    with tab3:
        st.markdown("## ⚙️ Gerenciamento de Categorias")
        if 'custom_categories' not in st.session_state:
            st.session_state.custom_categories = DEFAULT_CATEGORIAS_CONFIG.copy()
        else:
            for category_name, keywords in DEFAULT_CATEGORIAS_CONFIG.items():
                if category_name not in st.session_state.custom_categories:
                    st.session_state.custom_categories[category_name] = keywords
        analyzer = VendasAnalyzerWeb(st.session_state.custom_categories)
        st.markdown("Configure as categorias de produtos para uma análise mais precisa. Cada categoria tem palavras-chave que ajudam o sistema a identificar automaticamente os produtos.")
        
        st.markdown("### ➕ Adicionar Nova Categoria")
        col1, col2, col3 = st.columns([2, 3, 1])
        with col1: new_category_name = st.text_input("Nome da Categoria", placeholder="Ex: Camiseta Infantil", key="new_cat_name")
        with col2: new_keywords = st.text_input("Palavras-chave (separadas por vírgula)", placeholder="Ex: infantil, criança, baby", key="new_cat_keywords")
        with col3:
            if st.button("➕ Adicionar", use_container_width=True):
                if new_category_name.strip() and new_keywords.strip():
                    keywords_list = [kw.strip() for kw in new_keywords.split(',') if kw.strip()]
                    if analyzer.add_category(new_category_name.strip(), keywords_list):
                        st.session_state.custom_categories = analyzer.categorias_config.copy()
                        st.success(f"✅ Categoria '{new_category_name}' adicionada!")
                        st.rerun()
                    else: st.error("❌ Erro ao adicionar categoria. Verifique os dados.")
                else: st.error("❌ Preencha o nome da categoria e pelo menos uma palavra-chave.")
        
        st.markdown("---")
        st.markdown("### 📋 Categorias Configuradas")
        if not analyzer.get_categories_list():
            st.info("Nenhuma categoria configurada. Adicione uma acima.")
        else:
            for category_name in analyzer.get_categories_list():
                with st.expander(f"📦 {category_name}", expanded=False):
                    col1, col2, col3 = st.columns([2, 3, 1])
                    current_keywords = analyzer.categorias_config[category_name]
                    with col1: edit_name = st.text_input("Nome", value=category_name, key=f"edit_name_{category_name}")
                    with col2: edit_keywords = st.text_input("Palavras-chave", value=", ".join(current_keywords), key=f"edit_keywords_{category_name}")
                    with col3:
                        col3_1, col3_2 = st.columns(2)
                        with col3_1:
                            if st.button("💾 Salvar", key=f"save_{category_name}"):
                                new_keywords_list = [kw.strip() for kw in edit_keywords.split(',') if kw.strip()]
                                if analyzer.update_category(category_name, edit_name, new_keywords_list):
                                    st.session_state.custom_categories = analyzer.categorias_config.copy()
                                    st.success(f"✅ Categoria atualizada!")
                                    st.rerun()
                                else: st.error("❌ Erro ao atualizar categoria.")
                        with col3_2:
                            if st.button("🗑️ Remover", key=f"remove_{category_name}"):
                                if analyzer.remove_category(category_name):
                                    st.session_state.custom_categories = analyzer.categorias_config.copy()
                                    st.success(f"✅ Categoria '{category_name}' removida!")
                                    st.rerun()
                                else: st.error("❌ Erro ao remover categoria.")
            
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Resetar para Padrão", use_container_width=True):
                    st.session_state.custom_categories = DEFAULT_CATEGORIAS_CONFIG.copy()
                    st.success("✅ Categorias resetadas para configuração padrão!")
                    st.rerun()
            with col2:
                total_cats = len(analyzer.get_categories_list())
                total_keywords = sum(len(keywords) for keywords in analyzer.categorias_config.values())
                st.metric("Total de Categorias", total_cats)
                st.metric("Total de Palavras-chave", total_keywords)
        
        st.markdown("---")
        st.markdown("### 🧪 Teste de Categorização")
        test_product = st.text_input("Digite o nome de um produto para testar:", placeholder="Ex: Camiseta Oversized Preta")
        if test_product:
            detected_category = analyzer._identify_category(test_product)
            if detected_category == "Outros":
                st.warning(f"⚠️ Produto '{test_product}' foi categorizado como 'Outros'")
                st.info("💡 Adicione palavras-chave relevantes para categorizar este produto automaticamente.")
            else:
                st.success(f"✅ Produto '{test_product}' foi categorizado como: **{detected_category}**")
    
    with tab4:
        st.markdown("""
        ## 📖 Como Usar
        
        ### Passo 1: Prepare seu arquivo CSV
        - Exporte seus dados de vendas da Shopify ou plataforma de vendas
        - O arquivo deve conter as colunas: Name, Financial Status, Lineitem name, Lineitem quantity, Lineitem price, Subtotal, Shipping, Total, Payment Method
        - **Opcional:** Adicione uma coluna com informações de transportadora (Shipping Name, Shipping Method, Fulfillment Method, etc.) para análise de frete
        
        ### Passo 2: Carregue o arquivo
        - Clique em "Selecione seu arquivo CSV de vendas"
        - Escolha o arquivo da sua loja
        
        ### Passo 3: Configure os custos
        - Defina o custo de produção para cada categoria de produto
        - Configure o custo padrão para produtos não categorizados
        
        ### Passo 4: Configure as taxas por método de pagamento
        - **Taxas de Cartão**: Preencha as porcentagens dinâmicas por parcela (1x a 12x). O sistema buscará as parcelas automaticamente no CSV.
        - **Taxa Pix**: Porcentagem para vendas no Pix (geralmente 0%).
        - **Taxa Boleto**: Valor descontado por cada venda aprovada no boleto (Fixo em R$).
        - **Gateway e Antifraude**: Valor fixo (em R$) descontado para CADA pedido aprovado (Independente do método).
        - **Taxa Plataforma**: Taxa da plataforma (em %) cobrada sobre o volume de vendas.
        
        ### Passo 5: Configure custos de marketing
        - **Gasto com ADS**: Total gasto com publicidade (Google Ads, Facebook, TikTok, etc)
        - **Gestor de Tráfego**: Custo do profissional responsável pela gestão de campanhas
        
        ### Passo 6: Gere a análise
        - Clique em "GERAR ANÁLISE COMPLETA"
        - Visualize os resultados na tela
        - Baixe o relatório em PDF ou Excel
        """)

if __name__ == "__main__":
    if 'show_results' not in st.session_state:
        st.session_state.show_results = False
    if 'analysis_data' not in st.session_state:
        st.session_state.analysis_data = None
    
    main()
